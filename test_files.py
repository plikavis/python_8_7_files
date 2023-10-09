import os
from zipfile import ZipFile
from openpyxl import load_workbook
from pypdf import PdfReader
from xlrd import open_workbook

from examples.utils import RESOURCES_PATH


def test_check_sizes():
    with ZipFile('tmp/test.zip', 'r') as fz:
        for items in fz.infolist():
            file_name = items.filename
            src_size = os.path.getsize(f'resources/{file_name}')  # размер исходного файла
            dst_size = items.file_size  # размер файла в архиве
            assert src_size == dst_size, f'Размеры файлов c именем {file_name} не равны'


def test_check_txt():
    with ZipFile('tmp/test.zip', 'r') as fz:
        text = fz.read('text.txt ')  # файлы из архива
        with open('resources/text.txt ', 'r') as f:
            text_from_res = bytes(f.read(), 'ascii')
            assert text_from_res == text


def test_pdf():
    with ZipFile('tmp/test.zip', 'r') as fz:
        pdf_from_res = PdfReader(
            os.path.join(RESOURCES_PATH, "Python Testing with Pytest (Brian Okken).pdf"))  # pdf исходный
        pdf_from_arc = PdfReader(fz.open("Python Testing with Pytest (Brian Okken).pdf"))
        assert len(pdf_from_res.pages) == len(pdf_from_arc.pages)
        for i in range(len(pdf_from_res.pages)):
            assert pdf_from_res.pages[i].extract_text() == pdf_from_arc.pages[i].extract_text()


def test_xls():
    with (ZipFile('tmp/test.zip', 'r') as fz):
        xsl_from_res = open_workbook(os.path.join(RESOURCES_PATH, "file_example_XLS_10.xls"))
        xls_from_arc = open_workbook(file_contents=fz.read('file_example_XLS_10.xls'))
        assert xsl_from_res.nsheets == xls_from_arc.nsheets
        sheet_res = xsl_from_res.sheet_by_index(0)
        sheet_arc = xls_from_arc.sheet_by_index(0)
        for rx in range(sheet_res.nrows):
            for cx in range(sheet_res.ncols):
                assert sheet_res.cell_value(rx, cx) == sheet_arc.cell_value(rx, cx)


def test_xlsx():
    with (ZipFile('tmp/test.zip', 'r') as fz):
        workbook_res = load_workbook(os.path.join(RESOURCES_PATH, "file_example_XLSX_50.xlsx"))
        workbook_arc = load_workbook(fz.open('file_example_XLSX_50.xlsx'))
        sheet_res = workbook_res.active
        sheet_arc = workbook_arc.active
        assert sheet_arc.title == sheet_res.title
        rx = 1
        while rx <= sheet_res.max_row:
            rc = 1
            while rc <= sheet_res.max_column:
                assert sheet_res.cell(rx, rc).value == sheet_arc.cell(rx, rc).value
                rc += 1
            rx += 1
        workbook_res.close()
